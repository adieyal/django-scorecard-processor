from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from optparse import make_option
from openpyxl.reader.excel import load_workbook
from scorecard_processor import models
from django.utils import simplejson

class Command(BaseCommand):
    args = '<filename.xlsx>'
    help = 'Imports xls file into system'
    option_list = BaseCommand.option_list + (
        make_option('--name',
            dest='name',
            help='Give the survey a name'),
        )
    output_transaction = True


    #TODO: We can get this from option or database later
    SHEET_NAME = '1.1BirthRegistration'
    COLUM_NAME_ROW_STRING = "A3:H3"
    START_LINE = 3

    USER = User.objects.all()[1] #TODO: Fix this later


    def handle(self, *args, **options):
        try:
            wb = load_workbook(filename=args[0])
        except IndexError:
            raise CommandError("Require a filename")
        except :
            raise CommandError("Invalid file")


        survey_name = options.get('name')
        if not survey_name:
            raise CommandError("Require a name for the survey")


        project = models.Project.objects.get()
        sheet = wb.get_sheet_by_name(name=self.SHEET_NAME)

        column_names = {k:v.value for k, v in enumerate(sheet.range(self.COLUM_NAME_ROW_STRING)[0])}

        try:
            ds_group = models.DataSeriesGroup.objects.get(name=column_names[0])
            entity_type = models.EntityType.objects.get(name=column_names[0])
            del column_names[0]
        except models.DataSeriesGroup.DoesNotExist, models.EntityType.DoesNotExist:
            raise CommandError("Invalid file format")


        survey, created = models.Survey.objects.get_or_create(name=survey_name, project=project)

        if created:
            survey.data_series_groups.add(ds_group)
            survey.entity_types.add(entity_type)

        survey.question_set.all().delete()
        survey.questiongroup_set.all().delete()

        question = models.Question.objects.create(survey=survey, group=None,
            identifier=self.SHEET_NAME, question=self.SHEET_NAME)


        for i in xrange(self.START_LINE, sheet.get_highest_row()):
            cell = sheet.cell(row=i, column=0)

            entity, created = models.Entity.objects.get_or_create(name=cell.value,
                entity_type=entity_type, project=project)

            ds, created = models.DataSeries.objects.get_or_create(name=cell.value,
                group=ds_group)

            response_set, created = models.ResponseSet.objects.get_or_create(survey=survey,
                entity=entity)
            response_set.data_series.add(ds)

            value = {}
            for j in column_names:
                value[column_names[j]] = sheet.cell(row=i, column=j).value

            json = simplejson.dumps(value)

            response = models.Response.objects.create(question=question, response_set=response_set,
                respondant=self.USER, value=json)


        print "Done."

