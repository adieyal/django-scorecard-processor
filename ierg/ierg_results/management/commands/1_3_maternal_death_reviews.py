from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from optparse import make_option
from openpyxl.reader.excel import load_workbook
from scorecard_processor import models
from django.utils import simplejson

class Command(BaseCommand):
    args = '<filename.xlsx>'
    help = 'Imports xls file into system'
    option_list = BaseCommand.option_list + (
        make_option('--name',
            dest='name',
            help='Give the survey a name'),
        )
    output_transaction = True


    #TODO: We can get this from option or database later
    SHEET_NAME = '1.3MaternalDeathReviews'
    COLUM_COUNTRY = "A5"
    COLUM_COUNTRY_SR = 0
    COLUM_REGION = "B5"
    COLUM_REGION_SR = 1
    COLUM_YES_NO_FIRST = "C5"
    COLUM_YES_NO_FIRST_SR = 2
    COLUM_YES_NO_SECOND = "D5"
    COLUM_YES_NO_SECOND_SR = 3
    COLUM_QUESTIONS1 = "E5:G5"
    COLUM_QUESTIONS1_SR = 4
    COLUM_QUESTIONS2 = "H5:J5"
    COLUM_QUESTIONS2_SR = 7
    COLUM_QUESTIONS3 = "K5:M5"
    COLUM_QUESTIONS3_SR = 10
    COLUM_QUESTIONS4 = "N5:P5"
    COLUM_QUESTIONS4_SR = 13
    COLUM_QUESTIONS5 = "Q5:S5"
    COLUM_QUESTIONS5_SR = 16
    COLUM_QUESTIONS6 = "T5:V5"
    COLUM_QUESTIONS6_SR = 19
    START_LINE = 5

    USER = User.objects.all()[1] #TODO: Fix this later


    def handle(self, *args, **options):
        try:
            wb = load_workbook(filename=args[0])
        except IndexError:
            raise CommandError("Require a filename")
        except :
            raise CommandError("Invalid file")


        survey_name = options.get('name')
        if not survey_name:
            raise CommandError("Require a name for the survey")


        project = models.Project.objects.get()
        sheet = wb.get_sheet_by_name(name=self.SHEET_NAME)

        colum_questions1 = {k:v.value for k, v in\
            enumerate(sheet.range(self.COLUM_QUESTIONS1)[0], start=self.COLUM_QUESTIONS1_SR)}
        colum_questions2 = {k:v.value for k, v in\
            enumerate(sheet.range(self.COLUM_QUESTIONS2)[0], start=self.COLUM_QUESTIONS2_SR)}
        colum_questions3 = {k:v.value for k, v in\
            enumerate(sheet.range(self.COLUM_QUESTIONS3)[0], start=self.COLUM_QUESTIONS3_SR)}
        colum_questions4 = {k:v.value for k, v in\
            enumerate(sheet.range(self.COLUM_QUESTIONS4)[0], start=self.COLUM_QUESTIONS4_SR)}
        colum_questions5 = {k:v.value for k, v in\
            enumerate(sheet.range(self.COLUM_QUESTIONS5)[0], start=self.COLUM_QUESTIONS5_SR)}
        colum_questions6 = {k:v.value for k, v in\
            enumerate(sheet.range(self.COLUM_QUESTIONS6)[0], start=self.COLUM_QUESTIONS6_SR)}

        try:
            ds_group = models.DataSeriesGroup.objects.get(name=sheet.range(self.COLUM_COUNTRY).value)
            entity_type = models.EntityType.objects.get(name=sheet.range(self.COLUM_COUNTRY).value)
        except models.DataSeriesGroup.DoesNotExist, models.EntityType.DoesNotExist:
            raise CommandError("Invalid file format")


        survey, created = models.Survey.objects.get_or_create(name=survey_name, project=project)

        if created:
            survey.data_series_groups.add(ds_group)
            survey.entity_types.add(entity_type)

        survey.question_set.all().delete()
        survey.questiongroup_set.all().delete()

        question = models.Question.objects.create(survey=survey, group=None,
            identifier=self.SHEET_NAME, question=self.SHEET_NAME)


        for i in xrange(self.START_LINE, sheet.get_highest_row()):
            cell = sheet.cell(row=i, column=0)

            entity, created = models.Entity.objects.get_or_create(name=cell.value,
                entity_type=entity_type, project=project)

            ds, created = models.DataSeries.objects.get_or_create(name=cell.value,
                group=ds_group)

            response_set, created = models.ResponseSet.objects.get_or_create(survey=survey,
                entity=entity)
            response_set.data_series.add(ds)

            value = {}
            value[sheet.range(self.COLUM_REGION).value] =\
                sheet.cell(row=i, column=self.COLUM_REGION_SR).value
            value['data'] = []
            value['data'].append({sheet.range(self.COLUM_YES_NO_FIRST).value:\
                sheet.cell(row=i, column=self.COLUM_YES_NO_FIRST_SR).value})
            value['data'].append({sheet.range(self.COLUM_YES_NO_SECOND).value:\
                sheet.cell(row=i, column=self.COLUM_YES_NO_SECOND_SR).value})
            object_questions1 = {}
            for j in colum_questions1:
                object_questions1[colum_questions1[j]] = sheet.cell(row=i, column=j).value
            value['data'].append(object_questions1)
            object_questions2 = {}
            for j in colum_questions2:
                object_questions2[colum_questions2[j]] = sheet.cell(row=i, column=j).value
            value['data'].append(object_questions2)
            object_questions3 = {}
            for j in colum_questions3:
                object_questions3[colum_questions3[j]] = sheet.cell(row=i, column=j).value
            value['data'].append(object_questions3)
            object_questions4 = {}
            for j in colum_questions4:
                object_questions4[colum_questions4[j]] = sheet.cell(row=i, column=j).value
            value['data'].append(object_questions4)
            object_questions5 = {}
            for j in colum_questions5:
                object_questions5[colum_questions5[j]] = sheet.cell(row=i, column=j).value
            value['data'].append(object_questions5)
            object_questions6 = {}
            for j in colum_questions6:
                object_questions6[colum_questions6[j]] = sheet.cell(row=i, column=j).value
            value['data'].append(object_questions6)

            json = simplejson.dumps(value)

            response = models.Response.objects.create(question=question, response_set=response_set,
                respondant=self.USER, value=json)


        print "Done."

