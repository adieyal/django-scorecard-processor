from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from django.utils import simplejson
from openpyxl.reader.excel import load_workbook
from scorecard_processor import models
from ierg_results.models import ExcelFile
from optparse import make_option


class IergCommand(BaseCommand):
    args = '<filename.xlsx>'
    help = 'Imports xls file into system'
    option_list = BaseCommand.option_list + (
        make_option('--file_path',
            dest='file_path',
            help='Give the file path'),
        make_option('--excel_file_id',
            dest='excel_file_id',
            help='Give the excel file id'),
        )
    output_transaction = True

    #TODO: Fix this later
    USER = User.objects.all()[1]
    PROJECT = models.Project.objects.get()


    def handle(self, *args, **options):
        try:
            excel_file_id = options.get('excel_file_id')
            excel_file = ExcelFile.objects.get(id=excel_file_id)
        except Exception, e:
            raise CommandError("This file does not exist.")

        try:
            file_path = options.get('file_path')

            wb = load_workbook(filename=file_path)
            sheet = wb.get_sheet_by_name(name=self.SHEET_NAME)

            column_names = {k:v.value for k, v in\
                enumerate(sheet.range(self.COLUM_NAME_ROW_STRING)[0])}

            ds_group = models.DataSeriesGroup.objects.get(name=column_names[0])
            entity_type = models.EntityType.objects.get(name=column_names[0])
            del column_names[0]

            survey, created = models.Survey.objects.get_or_create(
                name=self.SURVEY_NAME, project=self.PROJECT)

            if created:
                survey.data_series_groups.add(ds_group)
                survey.entity_types.add(entity_type)

            question = models.Question.objects.create(survey=survey, group=None,
                identifier=self.IDENTIFIER, question=self.QUESTION)


            for i in xrange(self.START_LINE, self.FINISH_LINE):
                cell = sheet.cell(row=i, column=0)

                entity, created = models.Entity.objects.get_or_create(
                    name=cell.value, entity_type=entity_type, project=self.PROJECT)

                ds, created = models.DataSeries.objects.get_or_create(
                    name=cell.value, group=ds_group)

                response_set, created = models.ResponseSet.objects.get_or_create(
                    survey=survey, entity=entity)
                response_set.data_series.add(ds)

                json = self.get_json(sheet, column_names, i)

                response = models.Response.objects.create(question=question,
                    response_set=response_set, respondant=self.USER, value=json)


            excel_file.parse_log += '%s - parse completed\n' % self.IDENTIFIER
        except Exception, e:
            excel_file.parse_log += '%s - exception: %s\n' % (self.IDENTIFIER, e)
        finally:
            excel_file.save()


    def get_json(self, sheet, column_names, i):
        return simplejson.dumps({'json': None})

